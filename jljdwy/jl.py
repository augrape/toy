import openpyxl
from config.local_settings import logger
# 行以数字命名，列以字母命名；比如左上角第一个单元格的坐标为A1，下面为A2，右边为B1。
# openpyxl模块中有三个不同层次的类，workbook工作簿的抽象类，worksheet表格的抽象，cell单元格的抽象，每个类包含了许多属性和方法；
# 20190312110117237.xlsx

FILE_NAME = '8cdcd42dcac41df54b4708.xlsx'

# 塞选条件
PROVINCE = "吉林"
WORK = "工程"


def jilin_work_province(row):
    province = row[17].value
    work = row[3].value
    if PROVINCE in province and WORK == work:
        return [cell.value for cell in row]
        # print(*[cell.value for cell in row])


def created_jilin_work(wb, data, sheet_name='JL_province'):
    # is exist
    if sheet_name not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_name)
        logger.info(f" create {sheet_name} sheet.")
    else:
        logger.info(f"{sheet_name} sheet exist.")
        sheet = wb[sheet_name]
    row = 640
    for line in data:
        for col in range(1, len(line) + 1):
            sheet.cell(row=row, column=col).value = line[col - 1]
        row += 1

    wb.save(FILE_NAME)
    logger.info(f"{FILE_NAME} save.")


def main(wb, sheet_name):
    results = []
    logger.info(f"start to read {sheet_name} file.")
    for row in wb[sheet_name].iter_rows(min_row=7, max_row=15199,
                                        min_col=1, max_col=19):
        lst = jilin_work_province(row)
        if isinstance(lst, list):
            results.append(lst)
            logger.info(f"{lst}, append sucessfully.")

    created_jilin_work(wb, results)


if __name__ == '__main__':
    sheet_name = 'Sheet1'  # 全量表
    wb = openpyxl.load_workbook(FILE_NAME)
    main(wb, sheet_name)
